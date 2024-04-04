from flask import Flask, render_template,request,redirect,url_for,jsonify,send_from_directory,session,flash
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import extract,func,or_
from sqlalchemy.sql.expression import extract
from openpyxl import load_workbook
from datetime import date,datetime
from dateutil.parser import parse
from flask_migrate import Migrate
import logging
from werkzeug.utils import secure_filename
import os
import zipfile
import shutil
import subprocess
import json
from flask import current_app
app = Flask(__name__)
profile_images_upload_folder = 'static/profile_images'
if not os.path.exists(profile_images_upload_folder):
    os.makedirs(profile_images_upload_folder)

general_upload_folder = 'static/files'
if not os.path.exists(general_upload_folder):
    os.makedirs(general_upload_folder)    
app.config['SECRET_KEY'] = 'test'
app.config['SQLALCHEMY_DATABASE_URI'] = "sqlite:///employer.db"
app.config['SQLALCHEMY_BINDS']={'login':"sqlite:///login.db",
                                'delete_user':"sqlite:///delete.db",
                                'resume':"sqlite:///resume.db",
                                'intro':"sqlite:///intro.db",
                                'interview1':"sqlite:///interview1.db",
                                'interview2':"sqlite:///interview2.db",
                                'hr':"sqlite:///hr.db"
                                }

app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False
app.config['UPLOAD_FOLDER']=general_upload_folder
app.config['PROFILE_IMAGE_UPLOAD_FOLDER'] = 'static/profile_images'
db = SQLAlchemy(app)

class Employee(db.Model):
    Sno = db.Column(db.Integer, primary_key=True, autoincrement=True)
    Emp_id = db.Column(db.String(500))
    Name = db.Column(db.String(500))
    Designation = db.Column(db.String(500))
    Department = db.Column(db.String(500))
    Project = db.Column(db.String(500))
    Job_role = db.Column(db.String(500))
    Employment_status = db.Column(db.String(500))
    Joining_date = db.Column(db.String(500))
    Experience = db.Column(db.String(500))
    Location = db.Column(db.String(500))
    Last_promoted = db.Column(db.String(500))
    Comments = db.Column(db.String(500))
    employee_status=db.Column(db.String(500))


class Login(db.Model):
    __bind_key__="login"
    id=db.Column(db.Integer,primary_key=True)
    email=db.Column(db.String(500))
    password=db.Column(db.String(200))
    Role=db.Column(db.String(200))
    Name = db.Column(db.String(200))
    MobileNumber = db.Column(db.String(20))
    photo_filename = db.Column(db.String(255))
class Delete_user(db.Model):
    __bind_key__="delete_user"
    id=db.Column(db.Integer,primary_key=True) 
    Name=db.Column(db.String(200))
    Date=db.Column(db.String(200))   
class Resume(db.Model):
    __bind_key__="resume"  
    id=db.Column(db.Integer,primary_key=True)
    filename=db.Column(db.String(255), nullable=False)  

class Intro(db.Model):
    __bind_key__="intro"
    id=db.Column(db.Integer,primary_key=True)
    Date=db.Column(db.String(200))
    Status=db.Column(db.String(200))
    Comments=db.Column(db.String(200))
    resumeId = db.Column(db.Integer)  
    SelectedPanel=db.Column(db.String(200))  

class Interview1(db.Model):
    __bind_key__="interview1" 
    id=db.Column(db.Integer,primary_key=True)
    Date=db.Column(db.String(200))
    Status=db.Column(db.String(200))
    Comments=db.Column(db.String(200))
    resumeId=db.Column(db.Integer)  
    SelectedPanel=db.Column(db.String(200))     

class Interview2(db.Model):
    __bind_key__="interview2" 
    id=db.Column(db.Integer,primary_key=True)
    Date=db.Column(db.String(200))
    Status=db.Column(db.String(200))
    Comments=db.Column(db.String(200))
    resumeId=db.Column(db.Integer)  
    SelectedPanel=db.Column(db.String(200))    

class Hr(db.Model):
    __bind_key__="hr" 
    id=db.Column(db.Integer,primary_key=True)
    Date=db.Column(db.String(200))
    Status=db.Column(db.String(200))
    Comments=db.Column(db.String(200))
    resumeId=db.Column(db.Integer)  
    SelectedPanel=db.Column(db.String(200))        

def extract_data_from_excel():
    wb = load_workbook("employee_data 1.xlsx")
    ws = wb.active
    column_mappings = {
        'Sno': 0,
        'Emp_id': 1,
        'Name': 2,
        'Designation': 3,
        'Department': 4,
        'Project': 5,
        'Job_role': 6,
        'Employment_status': 7,
        'Joining_date': 8,
        'Experience': 9,
        'Location': 10,
        'Last_promoted': 11,
        'Comments': 12
    }
    for row in ws.iter_rows(min_row=2, values_only=True):
        
        if not all(cell is None for cell in row):
            Sno = row[column_mappings['Sno']]
            Emp_id = row[column_mappings['Emp_id']]
            Name = row[column_mappings['Name']]
            Designation = row[column_mappings['Designation']]
            Department = row[column_mappings['Department']]
            Project = row[column_mappings['Project']]
            Job_role = row[column_mappings['Job_role']]
            Employment_status = row[column_mappings['Employment_status']]
            Joining_date = row[column_mappings['Joining_date']]
            Experience = row[column_mappings['Experience']]
            formatted_date = None
            
            if isinstance(Joining_date, datetime):
                join_date = Joining_date 
                formatted_date=join_date.strftime("%d-%m-%Y")
                
                day = join_date.day
                month = join_date.month
                year = join_date.year

                current_date = date.today()
                Experience = current_date.year - year

                if (current_date.month, current_date.day) < (month, day):
                    Experience -= 1

                if Experience < 1:
                    Experience = "Less than 1 year"
            else:
                join_date = None
                day = None
                month = None
                year = None
                Experience = None

                 
            Location = row[column_mappings['Location']]
            Last_promoted = row[column_mappings['Last_promoted']]
            Comments = row[column_mappings['Comments']]
            


            existing_data = Employee.query.filter_by(Name=Name).first()
            if not existing_data:
                employee = Employee(Emp_id=Emp_id, Name=Name, Designation=Designation,
                                    Department=Department, Project=Project, Job_role=Job_role,
                                    Employment_status=Employment_status, Joining_date=formatted_date,
                                    Experience=Experience, Location=Location, Last_promoted=Last_promoted,
                                    Comments=Comments,employee_status="active")
                db.session.add(employee)
    db.session.commit()
    
    db.session.commit()
def allowed_file(filename):
    return "." in filename and filename.rsplit(".",1)[1] in ["xlsx","csv"]

def allowed_files(filename):
    return "." in filename and filename.rsplit(".",1)[1].lower() in ["pdf" , "csv","doc","docx"]

def dashboard_function():
    total_employees=Employee.query.count()
    active_employees = Employee.query.filter_by(employee_status='active').count()
    resigned_employees=Employee.query.filter_by(employee_status='resigned').count()
    total_resumes=Resume.query.count()
    hr_selected=Hr.query.filter_by(Status='Move to HR process').count()
    current_year = datetime.now().year
    last_year = current_year - 1
    
    
    current_year_count = Employee.query.filter(Employee.Joining_date.like(f'%-{current_year}')).count()
    
    
    last_year_count = Employee.query.filter(Employee.Joining_date.like(f'%-{last_year}')).count()
    Chennai_employees_count=Employee.query.filter_by(Location='Chennai').count()
    kollu_employees_count=Employee.query.filter_by(Location='Kollumangudi').count()
    kaup_employees_count=Employee.query.filter_by(Location='Kaup').count()
    tn_palyam_count=Employee.query.filter_by(Location='TN Palayam').count()
    return total_employees,active_employees,resigned_employees,total_resumes,hr_selected,current_year_count, last_year_count,Chennai_employees_count,kollu_employees_count,kaup_employees_count,tn_palyam_count
@app.context_processor
def inject_total_employees():
    total_employees,active_employees,resigned_employees,total_resumes,hr_selected,current_year_count, last_year_count,Chennai_employees_count,kollu_employees_count,kaup_employees_count,tn_palyam_count=dashboard_function()
    return dict(total_employees=total_employees,active_employees=active_employees,resigned_employees=resigned_employees,total_resumes=total_resumes,hr_selected=hr_selected,current_year_count=current_year_count, last_year_count=last_year_count,Chennai_employees_count=Chennai_employees_count,kollu_employees_count=kollu_employees_count,kaup_employees_count=kaup_employees_count,tn_palyam_count=tn_palyam_count)
@app.route("/",methods=["GET","POST"])
def signPage():
    correct_user=None
    error_message=None
    if request.method=="POST":
        email=request.form["email"]
        password=request.form["password"]
        
        correct_user=Login.query.filter_by(email=email).first()
        if correct_user:
            if correct_user.password==password:
                session['email'] = email 
                return redirect(url_for("dashBoard"))
            else:
                correct_user=None
                error_message="invalid login credentials"
        if correct_user == None:
                error_message="invalid login credentials"      
    return render_template("sign.html",error_message=error_message)   
@app.route("/dashboard")
def dashBoard():
    status=Employee.query.with_entities(Employee.Employment_status).distinct()
                   

    
    employment_status_counts={}
    for stat in status:
        count=Employee.query.filter_by(Employment_status=stat.Employment_status).count()
        
        employment_status_counts[stat.Employment_status]=count
    return render_template("dashboard.html",employment_status_counts=employment_status_counts) 
@app.route("/home",methods=["GET","POST"])
def Home():
    pages=20
    if request.method=="POST":
        pages=int(request.form["page"])
        print(pages)
            
    # data=Employee.query.all()
    # return render_template("index.html",data=data)
    page=request.args.get('page',1,type=int)
    data=Employee.query.paginate(page=page,per_page=pages)
    
    return render_template("index.html",data=data)
    
@app.route("/add", methods=["GET", "POST"])
def Add():
    if request.method == "POST":
        emp_id = request.form.get("emp_id")
        name = request.form.get("name")
        designation = request.form.get("designation")
        department = request.form.get("department")
        project = request.form.get("project")
        job_role = request.form.get("job_role")
        employment_status = request.form.get("employment_status")
        joining_date = request.form.get("joining_date")
        date_parts = joining_date.split('-')
        if len(date_parts) == 3:
            formatted_date = f"{date_parts[2]}-{date_parts[1]}-{date_parts[0]}"
            join_date=datetime.strptime(formatted_date, "%d-%m-%Y")
            current_date=date.today()   
            experience=current_date.year -int(join_date.year)
            if (current_date.month,current_date.day) < (join_date.month,join_date.day):
                experience-=1
            if experience <1:
                experience="Less than 1 year"       
        else:
            formatted_date = None 
            experience=None
        
        location = request.form.get("location")
        last_promoted = request.form.get("last_promoted")
        comments = request.form.get("comments")
        employee_status=request.form.get("status")
        existing_data=Employee.query.filter_by(Name=name).first()
        if existing_data:
            flash(f'Employee {name} already exists!', 'error')
        if not existing_data:
            employee = Employee(
                Emp_id=emp_id,
                Name=name,
                Designation=designation,
                Department=department,
                Project=project,
                Job_role=job_role,
                Employment_status=employment_status,
                Joining_date=formatted_date,
                Experience=experience,
                Location=location,
                Last_promoted=last_promoted,
                Comments=comments,
                employee_status=employee_status
            )
            db.session.add(employee)
            db.session.commit()
            flash(f'Added {name} successfully!', 'success')
        return redirect("/home")
    return render_template("add.html")

@app.route("/update/<int:sno>",methods=["GET","POST"])
def Update(sno):
    selected_date = request.args.get("date")
    if request.method == "POST":
        emp_id = request.form.get("emp_id")
        name = request.form.get("name")
        designation = request.form.get("designation")
        department = request.form.get("department")
        project = request.form.get("project")
        job_role = request.form.get("job_role")
        employment_status = request.form.get("employment_status")
        joining_date = request.form.get("joining_date")
        date_parts=joining_date.split("-")
        if len(date_parts)==3:
            formatted_date=f"{date_parts[2]}-{date_parts[1]}-{date_parts[0]}"
            join_date=datetime.strptime(formatted_date,"%d-%m-%Y")
            current_day=date.today()
            experience=current_day.year-int(join_date.year)
            if (current_day.month,current_day.day) < (join_date.month,join_date.day):
                experience-=1
            if experience < 1:
                experience="Less than 1 year"
        else:
            formatted_date = None 
            experience=None            
        
        location = request.form.get("location")
        last_promoted = request.form.get("last_promoted")
        comments = request.form.get("comments")
        employee_status=request.form.get("status")
        employee=Employee.query.filter_by(Sno=sno).first()
        employee.Emp_id=emp_id
        employee.Name=name
        employee.Designation=designation
        employee.Department=department
        employee.Project=project
        employee.Job_role=job_role
        employee.Employment_status=employment_status
        employee.Joining_date=formatted_date
        employee.Experience=experience
        employee.Location=location
        employee.Last_promoted=last_promoted
        employee.Comments=comments
        employee.employee_status=employee_status
        db.session.add(employee)
        db.session.commit()
        flash(f'{name} updated successfully!', 'success')
        return redirect("/home")
    employee=Employee.query.filter_by(Sno=sno).first()
    return render_template("update.html",employee=employee,selected_date=selected_date)

@app.route("/delete/<int:sno>")
def Delete(sno):
    employee=Employee.query.filter_by(Sno=sno).first()
    delete=Delete_user(Name=employee.Name,Date=employee.Joining_date)
    db.session.add(delete)
    db.session.commit()
    db.session.delete(employee)
    db.session.commit()
    return redirect("/home")
with app.app_context():
        db.create_all()
        # data=extract_data_from_excel()

@app.route("/bulk",methods=["GET","POST"])
def bulk():
    if request.method=="POST":
        file = request.files['file']
        if file and allowed_file(file.filename):
            if file.filename.endswith(".xlsx"):
                try:
                    wb=load_workbook(file)
                    ws=wb.active
                    column_mappings = {
                        'Sno': 0,
                        'Emp_id': 1,
                        'Name': 2,
                        'Designation': 3,
                        'Department': 4,
                        'Project': 5,
                        'Job_role': 6,
                        'Employment_status': 7,
                        'Joining_date': 8,
                        'Experience': 9,
                        'Location': 10,
                        'Last_promoted': 11,
                        'Comments': 12
                        }
                    for row in ws.iter_rows (min_row=2,values_only=True):
                        if not all(cell is None for cell in row):
                            
                            Emp_id = row[column_mappings['Emp_id']]
                            Name = row[column_mappings['Name']]
                            Designation = row[column_mappings['Designation']]
                            Department = row[column_mappings['Department']]
                            Project = row[column_mappings['Project']]
                            Job_role = row[column_mappings['Job_role']]
                            Employment_status = row[column_mappings['Employment_status']]
                            Joining_date = row[column_mappings['Joining_date']]
                            Experience = row[column_mappings['Experience']]
                            if Experience is None or Experience == "":
                                formatted_date = None
                                if isinstance(Joining_date, datetime):
                                    join_date = Joining_date
                                    formatted_date = join_date.strftime("%d-%m-%Y")
                                    month = join_date.month
                                    day = join_date.day
                                    year = join_date.year
                                    current_date = date.today()
                                    Experience = current_date.year - year
                                    if (current_date.month, current_date.day) < (month, day):
                                        Experience -= 1
                                    if Experience < 1:
                                        Experience = "Less than 1 year"
                                else:
                                    # If Joining_date is not a datetime object, set Experience to None
                                    Experience = None
                            else:
                                # If Experience is provided, keep formatted_date as is
                                formatted_date = Joining_date.strftime("%d-%m-%Y") if isinstance(Joining_date, datetime) else None
                    
                            Location = row[column_mappings['Location']]
                            Last_promoted = row[column_mappings['Last_promoted']]
                            Comments = row[column_mappings['Comments']]
                            employee_status="active"
                            existing_data=Employee.query.filter_by(Name=Name).first()
                            if existing_data:
                                flash( f'Employee {Name} aldready exists', 'error')
                            if not existing_data:
                                employee = Employee(Emp_id=Emp_id, Name=Name, Designation=Designation,
                                        Department=Department, Project=Project, Job_role=Job_role,
                                        Employment_status=Employment_status, Joining_date=formatted_date,
                                        Experience=Experience, Location=Location, Last_promoted=Last_promoted,
                                        Comments=Comments,employee_status=employee_status)
                                db.session.add(employee)
                                flash('File uploaded successfully!', 'success')
                    db.session.commit()
                    
                    return redirect("/home")            
                except Exception as e:
                    flash(f'Error: {e}', 'error')
                    return redirect("/bulk")
    
    return render_template("bulk.html")
    
@app.route("/view/<int:sno>")
def view(sno):
    data=Employee.query.filter_by(Sno=sno).first()
    return render_template("view.html",data=data)    
@app.route("/register",methods=["GET","POST"])
def register():
    if request.method=="POST":
        email=request.form["email"]
        password=request.form["password"]
        role = request.form["role"]
        name=request.form["name"]
        mobile=request.form["mobile"]
        filename=""
        if 'photo' in request.files:
            photo=request.files['photo']
            if photo.filename!='':
                filename=secure_filename(photo.filename)
                print(filename)
                file_path=os.path.join(app.config['PROFILE_IMAGE_UPLOAD_FOLDER'],filename)
                photo.save(file_path)
        user=Login(email=email,password=password,Role=role,Name=name,MobileNumber=mobile,photo_filename=filename)
        db.session.add(user)
        db.session.commit()
        return redirect("/")

    return render_template("register.html")
@app.route("/get_employees_list/<employment_status>")
def get_employees_list(employment_status):
    employees=Employee.query.filter_by(Employment_status=employment_status).all()
    return render_template('employee_list.html', employees=employees,employment_status=employment_status)
    # employee_names=[employee.Name for employee in employees]
    # return jsonify({'employeeList': employee_names})

@app.route("/resume",methods=["GET","POST"])
def resume():
    if request.method=="POST":
        selected_tag=request.form['tag']
        files=request.files.getlist('file')
        for file in files:
            if file and allowed_files(file.filename):
                filename=secure_filename(file.filename)
                new_filename=f"{selected_tag}_{filename}"
                target_path=os.path.join(app.config['UPLOAD_FOLDER'],new_filename)
                if not os.path.exists(target_path):
                    file.save(target_path)
                    flash('Resume uploaded successfully!', 'success')
                    resume=Resume(filename=new_filename)
                    db.session.add(resume)
                    db.session.commit()
                else:
                     flash(f'{filename} aldready exists !', 'error')   
        return redirect(url_for('resume'))        
    return render_template("resume.html")

@app.route("/employee_management")
def employee():
    page=request.args.get('page',1,type=int)
    resumes=Resume.query.paginate(page=page,per_page=10)
    return render_template("employee.html",resumes=resumes)

@app.route("/view_resume/<filename>")
def view_resume(filename):
    resume_folder='static/files'
    return send_from_directory(resume_folder,filename)

@app.route("/zip",methods=["GET","POST"])
def zip():
    if request.method=="POST":
        
        zip_files=request.files.getlist('zipFiles')
        selected_tag=request.form['tag']
        for zip_file in zip_files:
            if zip_file and zip_file.filename.endswith('.zip'):
                temp_dir="temp_dir"
                os.makedirs(temp_dir,exist_ok=True)
                try:
                    with zipfile.ZipFile(zip_file, 'r') as zip_ref:
                        zip_ref.extractall(temp_dir)
                    
                    for root,dirs,files in os.walk(temp_dir):
                        for filename in files:
                            file_path = os.path.join(root, filename)
                            if allowed_files(filename):
                                tag_filename=f"{selected_tag}_{filename}"
                                target_path=os.path.join(app.config['UPLOAD_FOLDER'], tag_filename)
                                print("File")
                                if not os.path.exists(target_path):
                                    shutil.move(file_path,target_path)
                                    resume=Resume(filename=tag_filename)
                                    db.session.add(resume)
                                    db.session.commit()
                                    flash('Resume uploaded successfully!', 'success')
                                else:
                                    flash('Resume already exists!', 'error')
                                    
                finally:
                    shutil.rmtree(temp_dir)   

        return redirect(url_for('employee'))            
    return render_template("zip.html")

@app.route("/introcall/<int:resume_id>", methods=['GET', 'POST'])
def introCall(resume_id):
    resume=Resume.query.get(resume_id)
    if request.method == 'POST':
        
        
        date=request.form["date"]
        status=request.form["status"]
        comments=request.form["comments"]
        selected_panel=request.form["selectedPanel"]
        existing_entry=Intro.query.filter_by(resumeId=resume.id).first()
        if existing_entry:
            existing_entry.Date = date
            existing_entry.Status = status
            existing_entry.Comments = comments
            existing_entry.SelectedPanel = selected_panel
        else:    
            entry=Intro( Date=date, Status=status, Comments=comments,resumeId=resume.id,SelectedPanel=selected_panel)
            db.session.add(entry)
        db.session.commit()
    return render_template("intro.html",resume=resume)

@app.route("/interview1")
def interview1():
    return render_template("interview1.html")

@app.route("/interview2")
def interview2():
    return render_template("interview2.html")

@app.route("/introv",methods=["GET","POST"])
def introv():
    if request.method=="POST":
        return render_template("introv.html")

@app.route("/resume_details/<int:resume_id>")
def resume_details(resume_id):
    resume=Resume.query.get(resume_id)
    intro_call=Intro.query.filter_by(resumeId=resume_id).first()
    interview1=Interview1.query.filter_by(resumeId=resume_id).first()
    interview2=Interview2.query.filter_by(resumeId=resume_id).first()
    hr=Hr.query.filter_by(resumeId=resume_id).first()
    return render_template("resume_details.html",resume=resume,intro_call=intro_call,interview1=interview1,interview2=interview2,hr=hr)

@app.route("/interview1v/<int:resume_id>", methods=["GET", "POST"])
def interview1v(resume_id):
    resume=Resume.query.get(resume_id)
    if request.method=="POST":
        date=request.form["date"]
        comments=request.form["comments"]
        status=request.form["status"]
        
        selected_panel=request.form["selectedPanel"]
        existing_entry=Interview1.query.filter_by(resumeId=resume.id).first()
        if existing_entry:
            existing_entry.Date = date
            existing_entry.Status = status
            existing_entry.Comments = comments
            existing_entry.SelectedPanel = selected_panel
        else:    
            entry=Interview1(Date=date, Status=status, Comments=comments,resumeId=resume.id,SelectedPanel=selected_panel)
            db.session.add(entry)
        db.session.commit()
    resume = Resume.query.get(resume_id)
    return render_template("interview1.html", resume=resume)

@app.route("/interview2v/<int:resume_id>",methods=["GET", "POST"])
def interview2v(resume_id):
    resume=Resume.query.get(resume_id)
    if request.method=="POST":
        
        date=request.form["date"]
        comments=request.form["comments"]
        status=request.form["status"]
        selected_panel=request.form["selectedPanel"]
        existing_entry=Interview2.query.filter_by(resumeId=resume.id).first()
        if existing_entry:
            existing_entry.Date = date
            existing_entry.Status = status
            existing_entry.Comments = comments
            existing_entry.SelectedPanel = selected_panel
        else:    
            entry=Interview2(Date=date, Status=status, Comments=comments,resumeId=resume.id,SelectedPanel=selected_panel)
            db.session.add(entry)
        db.session.commit()
    return render_template("interview2.html",resume=resume)

@app.route("/hr/<int:resume_id>",methods=["GET", "POST"])
def hr(resume_id):
    resume=Resume.query.get(resume_id)
    if request.method=="POST":
        date=request.form["date"]
        comments=request.form["comments"]
        status=request.form["status"]
        selected_panel=request.form["selectedPanel"]
        existing_entry=Hr.query.filter_by(resumeId=resume.id).first()
        if existing_entry:
            existing_entry.Date = date
            existing_entry.Status = status
            existing_entry.Comments = comments
            existing_entry.SelectedPanel = selected_panel
        else:     
            entry=Hr(Date=date, Status=status, Comments=comments,resumeId=resume.id,SelectedPanel=selected_panel)
            db.session.add(entry)
        db.session.commit()
    return render_template("hr.html",resume=resume)

@app.route("/get_interview_status/<int:resume_id>")
def get_intro_status(resume_id):
    intro=Intro.query.filter_by(resumeId=resume_id).first()
    interview1=Interview1.query.filter_by(resumeId=resume_id).first()
    interview2=Interview2.query.filter_by(resumeId=resume_id).first()
    hr=Hr.query.filter_by(resumeId=resume_id).first()
    intro_status=intro.Status if intro else "Intro call not conducted"
    interview1_status=interview1.Status if interview1 else "Interview 1 not conducted"
    interview2_status=interview2.Status if interview2 else "Interview 2 not conducted"
    hr_status=hr.Status if hr else "Hr not conducted"
    all_rounds_status="cleared" if all([intro, interview1, interview2, hr]) else "Not cleared"
    return jsonify({'intro_status':intro_status,'interview1_status': interview1_status,'interview2_status':interview2_status,'hr_status':hr_status, 'all_rounds_status': all_rounds_status}) 

@app.route("/employee_data")
def employeeData():
    data=Employee.query.all()
    return render_template("employee_data.html",data=data)

@app.route("/profile")
def profile():
    email = session.get('email')
    if email:
        user=Login.query.filter_by(email=email).first()
        emailId=user.email
        password=user.password
        role=user.Role
        mobile=user.MobileNumber
        name=user.Name
        filename = user.photo_filename
        return render_template("profile.html",emailId=emailId,password=password,role=role,name=name,mobile=mobile,filename=filename)

@app.route("/get_role")
def get_role():
    email = session.get('email')
    if email:
        user=Login.query.filter_by(email=email).first()
        role=user.Role
        return jsonify({'role': role})
@app.route("/signout")
def signout():
    session.pop('email', None)    
    return redirect(url_for('signPage'))
    


@app.route('/edit_config')
def edit_config():
    return render_template("config_editor.html")
    
    # subprocess.run(['notepad', 'static/config.js'])  
    # return 'File edited successfully'  

@app.route("/update_profile", methods=["POST"])
def update_profile():
    email = session.get('email')
    
    if email:
        user = Login.query.filter_by(email=email).first()
        user.Name = request.form["name"]
        
        user.MobileNumber = request.form["mobile"]
        user.password = request.form["password"]
        # user.Role = request.form.get("role") 
        
        if 'photo' in request.files:
            photo = request.files['photo']
            if photo.filename != '':
                # Delete previous photo if it exists
                if user.photo_filename:
                    previous_photo_path = os.path.join(app.config['PROFILE_IMAGE_UPLOAD_FOLDER'], user.photo_filename)
                    if os.path.exists(previous_photo_path):
                        os.remove(previous_photo_path)
                # Save new photo
                filename = secure_filename(photo.filename)
                file_path = os.path.join(app.config['PROFILE_IMAGE_UPLOAD_FOLDER'], filename)
                photo.save(file_path)
                user.photo_filename = filename
        db.session.commit()
        return redirect("/profile")
         
@app.route('/update_config', methods=['POST'])
def update_config():
    try:
        # Get the updated config data from the request
        updated_config = request.get_json()

        # Write the updated config data to the config.js file
        with open('static/config.js', 'w') as config_file:
            config_file.write(f'var config = {json.dumps(updated_config)};')

        return jsonify({'message': 'Config updated successfully'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500    

if __name__ == "__main__":
    app.run(debug=True)



